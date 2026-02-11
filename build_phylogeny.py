#!/usr/bin/env python3
import argparse
import os
import re
import subprocess
import shutil
from openpyxl import load_workbook

GROUP_ORDER = ["CALCR", "CRHR", "GCGR", "SCTR", "PTHR"]


def find_latest_run(outputs_root):
    if not os.path.isdir(outputs_root):
        return None
    runs = [d for d in os.listdir(outputs_root) if d.startswith("run_")]
    if not runs:
        return None
    runs.sort()
    return os.path.join(outputs_root, runs[-1])


def sanitize_label(value):
    safe = re.sub(r"[^A-Za-z0-9_]+", "_", value).strip("_")
    return safe or "seq"


def normalize_for_match(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value).lower()).strip()


def classify_from_description(description_line):
    desc_norm = normalize_for_match(description_line)
    crhr2_keys = [
        "crhr2",
        "crfr2",
        "crh receptor 2",
        "crf receptor 2",
        "corticotropin releasing factor receptor 2",
        "corticotropin releasing hormone receptor 2",
    ]
    crhr1_keys = [
        "crhr1",
        "crfr1",
        "crh receptor 1",
        "crf receptor 1",
        "corticotropin releasing factor receptor 1",
        "corticotropin releasing hormone receptor 1",
    ]
    if any(k in desc_norm for k in crhr2_keys):
        return "CRHR2", "CRHR"
    if any(k in desc_norm for k in crhr1_keys):
        return "CRHR1", "CRHR"
    return None, None


def collect_all_sequences(excel_path):
    wb = load_workbook(excel_path, read_only=True)
    rows = []
    for sheet_name in list(GROUP_ORDER) + ["UNCLASSIFIED"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_skipped = False
        for row in ws.iter_rows(values_only=True):
            if not header_skipped:
                header_skipped = True
                continue
            if not row:
                continue
            if str(row[0]).strip().upper() == "REJEITADAS ABAIXO":
                break
            subfamily = row[0]
            accession = row[1]
            sequence = row[6]
            species = row[7]
            description = row[8] if len(row) > 8 else ""
            family = sheet_name if sheet_name in GROUP_ORDER else None

            if sheet_name == "UNCLASSIFIED":
                subfamily, family = classify_from_description(description or "")
                if not family:
                    continue

            if not accession or not sequence:
                continue
            label = (
                f"{sanitize_label(species)}_"
                f"{sanitize_label(accession)}_"
                f"{sanitize_label(family)}_"
                f"{sanitize_label(subfamily)}"
            )
            rows.append((label, str(sequence).replace(" ", "")))
    return rows


def write_fasta(records, output_path):
    with open(output_path, "w") as f:
        for label, seq in records:
            f.write(f">{label}\n{seq}\n")


def run_cmd(cmd, cwd=None):
    subprocess.run(cmd, check=True, cwd=cwd)


def resolve_fasttree_cmd():
    for candidate in ["fasttree", "FastTree"]:
        if shutil.which(candidate):
            return candidate
    raise FileNotFoundError("fasttree not found in PATH")


def parse_args():
    parser = argparse.ArgumentParser(
        description="Build phylogeny from per-species Excel outputs"
    )
    parser.add_argument(
        "--run-root",
        default=None,
        help="Path to a specific outputs/run_* folder (default: latest run)",
    )
    parser.add_argument(
        "--mafft-mode",
        default="quick",
        choices=["quick", "auto", "linsi", "ginsi", "einsi"],
        help="Alignment strategy: quick (fast) or high-accuracy presets",
    )
    parser.add_argument(
        "--prefix",
        default="ALL_CLASSIFIED",
        help="Output prefix for fasta/aln/tree files",
    )
    return parser.parse_args()


def build_mafft_cmd(mode, fasta_path):
    if mode == "quick":
        return [
            "mafft",
            "--retree",
            "2",
            "--maxiterate",
            "2",
            fasta_path,
        ]
    if mode == "auto":
        return ["mafft", "--auto", fasta_path]
    if mode == "linsi":
        return ["mafft", "--localpair", "--maxiterate", "1000", fasta_path]
    if mode == "ginsi":
        return ["mafft", "--globalpair", "--maxiterate", "1000", fasta_path]
    if mode == "einsi":
        return ["mafft", "--genafpair", "--maxiterate", "1000", fasta_path]
    raise ValueError(f"Unsupported MAFFT mode: {mode}")


def main():
    args = parse_args()
    workspace_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    outputs_root = os.path.join(workspace_root, "outputs")
    run_root = args.run_root or find_latest_run(outputs_root)
    if not run_root:
        raise SystemExit("No outputs/run_* folder found.")
    run_root = os.path.abspath(run_root)

    phylo_root = os.path.join(run_root, "phylogeny_all")
    os.makedirs(phylo_root, exist_ok=True)

    excel_files = [
        os.path.join(run_root, f)
        for f in os.listdir(run_root)
        if f.endswith("_FamilyB1.xlsx")
    ]
    if not excel_files:
        raise SystemExit("No per-species Excel files found in latest run.")

    all_records = []
    for excel_path in excel_files:
        all_records.extend(collect_all_sequences(excel_path))

    if not all_records:
        raise SystemExit("No classified sequences found.")

    fasta_path = os.path.join(phylo_root, f"{args.prefix}.fasta")
    aln_path = os.path.join(phylo_root, f"{args.prefix}.aln.fasta")
    write_fasta(all_records, fasta_path)

    mafft_cmd = build_mafft_cmd(args.mafft_mode, fasta_path)
    print(f"MAFFT mode: {args.mafft_mode}")
    with open(aln_path, "w") as out_f:
        subprocess.run(mafft_cmd, check=True, stdout=out_f)

    tree_path = os.path.join(phylo_root, f"{args.prefix}.fasttree.nwk")
    fasttree_cmd = resolve_fasttree_cmd()
    with open(tree_path, "w") as out_f:
        subprocess.run([fasttree_cmd, "-lg", "-gamma", aln_path], check=True, stdout=out_f)

    print(f"Phylogeny outputs saved to: {phylo_root}")
    print(f"FastTree tree: {tree_path}")


if __name__ == "__main__":
    main()
