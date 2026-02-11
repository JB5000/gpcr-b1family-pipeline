#!/usr/bin/env python3
import argparse
import json
import os
import re
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook

FAMILIES = {
    "SCTR": ["secretin receptor", "secretin", "sctr", "sctra", "sctrb"],
    "GCGR": ["glucagon receptor", "gcgr", "gcgra", "gcgrb", "gcrpr"],
    "GIPR": ["gastric inhibitory", "gipr", "gip receptor", "gipr2"],
    "GLP1R": [
        "glucagon-like peptide-1",
        "glucagon like peptide 1",
        "glucagon-like peptide 1 receptor",
        "glp1r",
        "glp-1 receptor",
        "glpr1",
    ],
    "GLP2R": [
        "glucagon-like peptide-2",
        "glucagon like peptide 2",
        "glucagon-like peptide 2 receptor",
        "glp2r",
        "glp-2 receptor",
        "glpr2",
    ],
    "GHRHR": [
        "growth hormone releasing hormone receptor",
        "growth hormone releasing hormone receptor, like",
        "growth hormone-releasing",
        "ghrhr",
    ],
    "PTH1R": [
        "parathyroid hormone receptor 1",
        "parathyroid hormone/parathyroid hormone-related peptide receptor",
        "parathyroid hormone-related peptide receptor",
        "parathyroid hormone/parathyroid hormone-related peptide receptor-like",
        "pth1r",
        "pthr1",
        "pthr1r",
        "pth3r",
        "pthr3",
        "prpr",
        "prpra",
        "prprb",
        "phir",
        "pihr",
    ],
    "PTH2R": [
        "parathyroid hormone receptor 2",
        "parathyroid hormone 2 receptor",
        "pth2r",
        "pth2ra",
        "pth2rb",
        "pthr2",
        "pthr2r",
    ],
    "CRHR1": [
        "crhr1",
        "crfr1",
        "crh receptor 1",
        "crh-r1",
        "crf receptor 1",
        "crf1 receptor",
        "corticotropin releasing factor receptor 1",
        "corticotropin releasing hormone receptor 1",
        "corticotropin-releasing hormone receptor 1",
    ],
    "CRHR2": [
        "crhr2",
        "crfr2",
        "crh receptor 2",
        "crh-r2",
        "crf receptor 2",
        "crf2 receptor",
        "corticotropin releasing factor receptor 2",
        "corticotropin releasing hormone receptor 2",
        "corticotropin-releasing hormone receptor 2",
    ],
    "ADCYAP1R1": [
        "pacap",
        "adcyap1r1",
        "adcyap1r1a",
        "adcyap1r1b",
        "pituitary adenylate cyclase-activating polypeptide type i receptor",
        "pituitary adenylate cyclase-activating polypeptide type 1 receptor",
        "adenylate cyclase activating polypeptide 1",
        "pac1 receptor",
        "pacap type i receptor",
    ],
    "VIPR1": [
        "vasoactive intestinal peptide receptor 1",
        "vasoactive intestinal polypeptide receptor 1",
        "vipr1",
        "vipr1a",
        "vipr1b",
    ],
    "VIPR2": [
        "vasoactive intestinal peptide receptor 2",
        "vasoactive intestinal polypeptide receptor 2",
        "vipr2",
        "vipr2a",
        "vipr2b",
    ],
    "CALCR": ["calcitonin receptor", "calcr", "calcra", "calcrb", "calcr3"],
    "CALCRL": [
        "calcitonin receptor-like",
        "calcitonin gene-related peptide type 1 receptor",
        "calcitonin gene-related peptide type 1 receptor-like",
        "calcitonin gene related peptide type 1 receptor",
        "calcrl",
        "calcrla",
        "calcrlb",
        "calcrl3",
    ],
}

CLASSIFICATION_ORDER = [
    "VIPR2",
    "VIPR1",
    "GLP2R",
    "GLP1R",
    "PTH2R",
    "PTH1R",
    "CRHR2",
    "CRHR1",
    "SCTR",
    "GCGR",
    "GIPR",
    "GHRHR",
    "ADCYAP1R1",
    "CALCR",
    "CALCRL",
]

GROUP_ORDER = ["CALCR", "CRHR", "GCGR", "SCTR", "PTHR"]

GROUP_SUBFAMILIES = {
    "CALCR": ["CALCRL", "CALCR"],
    "CRHR": ["CRHR1", "CRHR2"],
    "GCGR": ["GCGR", "GLP1R", "GLP2R", "GIPR"],
    "SCTR": ["GHRHR", "ADCYAP1R1", "VIPR1", "VIPR2", "SCTR"],
    "PTHR": ["PTH1R", "PTH2R"],
}

SUBFAMILY_TO_GROUP = {
    "CALCRL": "CALCR",
    "CALCR": "CALCR",
    "CRHR1": "CRHR",
    "CRHR2": "CRHR",
    "GCGR": "GCGR",
    "GLP1R": "GCGR",
    "GLP2R": "GCGR",
    "GIPR": "GCGR",
    "GHRHR": "SCTR",
    "ADCYAP1R1": "SCTR",
    "VIPR1": "SCTR",
    "VIPR2": "SCTR",
    "SCTR": "SCTR",
    "PTH1R": "PTHR",
    "PTH2R": "PTHR",
}


def parse_args():
    parser = argparse.ArgumentParser(description="Build per-species/combined Excel files from cache JSON")
    parser.add_argument("--cache", required=True, help="Path to JSON cache from extract_ncbi_cache.py")
    parser.add_argument(
        "--output-root",
        default=None,
        help="Output folder (default: outputs/run_<timestamp>)",
    )
    return parser.parse_args()


def sanitize_name(value):
    safe = re.sub(r"[^A-Za-z0-9]+", "_", str(value)).strip("_")
    return safe or "output"


def normalize_for_match(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value).lower()).strip()


NORMALIZED_FAMILY_KEYS = {
    subfamily: [normalize_for_match(k) for k in keys]
    for subfamily, keys in FAMILIES.items()
}


def classify_subfamily(description_line):
    desc_norm = normalize_for_match(description_line)
    for subfamily in CLASSIFICATION_ORDER:
        keys = NORMALIZED_FAMILY_KEYS.get(subfamily, [])
        if any(k and k in desc_norm for k in keys):
            return subfamily
    return "UNCLASSIFIED"


def get_species_abbrev(description_line):
    try:
        match = re.findall(r"\[([^\]]+)\]", description_line)
        if not match:
            return "NA"
        species = match[-1].strip()
        parts = [p for p in species.split() if p]
        if len(parts) < 2:
            return "NA"
        return f"{parts[0][0].upper()}{parts[-1][:2].lower()}"
    except Exception:
        return "NA"


def is_np(accession):
    return str(accession).startswith("NP_")


def build_row(record, subfamily, family, accession_data, gene_summary):
    acc = record["accession"]
    meta = accession_data.get(acc, {})
    gene_id = meta.get("gene_id") or "NA"
    ginfo = gene_summary.get(gene_id, {}) if gene_id != "NA" else {}
    return {
        "order": record["line_number"],
        "family": family,
        "subfamily": subfamily,
        "accession": acc,
        "tlen": record.get("tlen") or "NA",
        "gene_id": gene_id,
        "chromosome": ginfo.get("chromosome") or "NA",
        "exon_count": ginfo.get("exon_count") or "NA",
        "sequence": meta.get("sequence") or "NA",
        "species": get_species_abbrev(record["line"]),
        "description": record["line"],
    }


def write_phmmer_sheet(wb, input_file, raw_records):
    ws = wb.active
    ws.title = "PHMMER_RESULTS"
    ws.append(["source_file", "line"])
    for rec in raw_records:
        ws.append([os.path.basename(input_file), rec["line"]])


def write_family_sheet(wb, family, selected_rows, rejected_rows):
    ws = wb.create_sheet(title=family)
    ws.append([
        "Subfamily",
        "Accession",
        "tlen",
        "GeneID",
        "chromosome",
        "exon_count",
        "Sequence",
        "Species",
        "Description",
    ])

    for row in selected_rows:
        ws.append([
            row["subfamily"],
            row["accession"],
            row["tlen"],
            row["gene_id"],
            row["chromosome"],
            row["exon_count"],
            row["sequence"],
            row["species"],
            row["description"],
        ])

    if rejected_rows:
        ws.append(["REJEITADAS ABAIXO", None, None, None, None, None, None, None, None])
        for row in rejected_rows:
            ws.append([
                row["subfamily"],
                row["accession"],
                row["tlen"],
                row["gene_id"],
                row["chromosome"],
                row["exon_count"],
                row["sequence"],
                row["species"],
                row["description"],
            ])


def write_species_excel(species_data, output_root):
    wb = Workbook()
    write_phmmer_sheet(wb, species_data["input_file"], species_data["raw_records"])

    selected = species_data["selected"]
    rejected = species_data["rejected"]

    for family in GROUP_ORDER:
        selected_rows = [r for r in selected if r["family"] == family]
        rejected_rows = [r for r in rejected if r["family"] == family]
        write_family_sheet(wb, family, selected_rows, rejected_rows)

    unclassified_selected = [r for r in selected if r["family"] == "UNCLASSIFIED"]
    unclassified_rejected = [r for r in rejected if r["family"] == "UNCLASSIFIED"]
    write_family_sheet(wb, "UNCLASSIFIED", unclassified_selected, unclassified_rejected)

    out_path = os.path.join(output_root, f"{species_data['safe_base']}_FamilyB1.xlsx")
    wb.save(out_path)
    return out_path


def normalize_sheet_title(value, used_titles):
    base = sanitize_name(value)[:31] or "Sheet"
    title = base
    suffix = 1
    while title in used_titles:
        title = f"{base[:28]}_{suffix}"
        suffix += 1
    used_titles.add(title)
    return title


def write_combined_workbook(all_species_data, output_root):
    wb = Workbook()
    wb.remove(wb.active)
    used_titles = set()

    for data in all_species_data:
        sheet_title = normalize_sheet_title(data["safe_base"], used_titles)
        ws = wb.create_sheet(title=sheet_title)
        ws.append([
            "Family",
            "Subfamily",
            "Accession",
            "tlen",
            "GeneID",
            "chromosome",
            "exon_count",
            "Sequence",
            "Species",
            "Description",
        ])

        selected = data["selected"]
        rejected = data["rejected"]

        for family in GROUP_ORDER:
            for subfamily in GROUP_SUBFAMILIES[family]:
                rows = [r for r in selected if r["subfamily"] == subfamily]
                if rows:
                    for row in rows:
                        ws.append([
                            family,
                            subfamily,
                            row["accession"],
                            row["tlen"],
                            row["gene_id"],
                            row["chromosome"],
                            row["exon_count"],
                            row["sequence"],
                            row["species"],
                            row["description"],
                        ])
                else:
                    ws.append([family, subfamily, None, None, None, None, None, None, None, None])

        unclassified_selected = [r for r in selected if r["family"] == "UNCLASSIFIED"]
        if unclassified_selected:
            for row in unclassified_selected:
                ws.append([
                    "UNCLASSIFIED",
                    "UNCLASSIFIED",
                    row["accession"],
                    row["tlen"],
                    row["gene_id"],
                    row["chromosome"],
                    row["exon_count"],
                    row["sequence"],
                    row["species"],
                    row["description"],
                ])
        else:
            ws.append(["UNCLASSIFIED", "UNCLASSIFIED", None, None, None, None, None, None, None, None])

        if rejected:
            ws.append(["REJEITADAS ABAIXO", None, None, None, None, None, None, None, None, None])
            for row in rejected:
                ws.append([
                    row["family"],
                    row["subfamily"],
                    row["accession"],
                    row["tlen"],
                    row["gene_id"],
                    row["chromosome"],
                    row["exon_count"],
                    row["sequence"],
                    row["species"],
                    row["description"],
                ])

    out_path = os.path.join(output_root, "All_species_combined.xlsx")
    wb.save(out_path)
    return out_path


def run_for_species(species_block, accession_data, gene_summary):
    records = species_block.get("records", [])

    by_gene = defaultdict(list)
    for record in records:
        acc = record["accession"]
        gene_id = accession_data.get(acc, {}).get("gene_id") or "NA"
        key = gene_id if gene_id != "NA" else f"NO_GENE_{acc}"
        by_gene[key].append(record)

    selected = []
    eliminated = []

    for _, group in by_gene.items():
        group.sort(key=lambda x: ((x.get("tlen") or -1), is_np(x["accession"])), reverse=True)
        selected.append(group[0])
        eliminated.extend(group[1:])

    selected_rows = []
    rejected_rows = []

    for r in sorted(selected, key=lambda x: x["line_number"]):
        subfamily = classify_subfamily(r["line"])
        family = SUBFAMILY_TO_GROUP.get(subfamily, "UNCLASSIFIED")
        if subfamily == "UNCLASSIFIED":
            family = "UNCLASSIFIED"
        selected_rows.append(build_row(r, subfamily, family, accession_data, gene_summary))

    for r in sorted(eliminated, key=lambda x: x["line_number"]):
        subfamily = classify_subfamily(r["line"])
        family = SUBFAMILY_TO_GROUP.get(subfamily, "UNCLASSIFIED")
        if subfamily == "UNCLASSIFIED":
            family = "UNCLASSIFIED"
        rejected_rows.append(build_row(r, subfamily, family, accession_data, gene_summary))

    return selected_rows, rejected_rows


def resolve_output_root(path):
    if path:
        return path
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(os.getcwd(), "outputs", f"run_{run_id}")


def main():
    args = parse_args()
    with open(args.cache) as f:
        cache = json.load(f)

    output_root = resolve_output_root(args.output_root)
    os.makedirs(output_root, exist_ok=True)

    accession_data = cache.get("accession_data", {})
    gene_summary = cache.get("gene_summary", {})
    file_records = cache.get("file_records", {})

    all_species = []
    for safe_base, block in file_records.items():
        selected, rejected = run_for_species(block, accession_data, gene_summary)
        species_data = {
            "input_file": block.get("input_file", safe_base),
            "safe_base": safe_base,
            "raw_records": block.get("records", []),
            "selected": selected,
            "rejected": rejected,
        }
        all_species.append(species_data)
        out = write_species_excel(species_data, output_root)
        print(f"Excel saved: {out}")

    combined = write_combined_workbook(all_species, output_root)
    print(f"Combined Excel saved: {combined}")
    print(f"Output root: {output_root}")


if __name__ == "__main__":
    main()
