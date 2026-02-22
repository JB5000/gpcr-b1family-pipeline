#!/usr/bin/env python3
"""
Generate a combined Excel workbook with 5 main family sheets plus PHMMER results.
The layout follows FamilyB1_Seqs_final.xlsx (grouped subfamilies).
"""
import os
import re
import csv
import argparse
from openpyxl import Workbook

DEFAULT_OUTPUT_DIR = "FINAL_OUTPUT"

GROUP_ORDER = ["CALCR", "CRHR", "GCGR", "SCTR", "PTHR"]
GROUP_SUBFAMILIES = {
    "CALCR": ["CALCRL", "CALCR"],
    "CRHR": ["CRHR1", "CRHR2"],
    "GCGR": ["GCGR", "GLP1R", "GLP2R", "GIPR"],
    "SCTR": ["GHRHR", "ADCYAP1R1", "VIPR1", "VIPR2", "SCTR"],
    "PTHR": ["PTH1R", "PTH2R"],
}


def sanitize_name(value):
    safe = re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_")
    return safe or "output"


def resolve_output_dir(input_file, output_root=None):
    base = os.path.splitext(os.path.basename(input_file))[0]
    safe_base = sanitize_name(base)
    output_dir = f"FINAL_OUTPUT_{safe_base}"
    if output_root:
        return os.path.join(output_root, output_dir)
    return output_dir


def find_latest_output_root():
    base = os.path.join(os.getcwd(), "outputs")
    if not os.path.isdir(base):
        return None
    candidates = []
    for name in os.listdir(base):
        if name.startswith("run_"):
            candidates.append(os.path.join(base, name))
    if not candidates:
        return None
    return sorted(candidates)[-1]


def parse_args():
    parser = argparse.ArgumentParser(
        description="Generate combined Excel output for GPCR B1 pipeline"
    )
    parser.add_argument(
        "-i",
        "--input",
        action="append",
        dest="inputs",
        default=[],
        help="PHMMER input file used in the pipeline (can be used multiple times)",
    )
    parser.add_argument(
        "-O",
        "--output-dir",
        action="append",
        dest="output_dirs",
        default=[],
        help="Pipeline output directory (can be used multiple times)",
    )
    parser.add_argument(
        "--output-root",
        dest="output_root",
        default=None,
        help="Folder containing multiple FINAL_OUTPUT_* directories",
    )
    parser.add_argument(
        "--phmmer-dir",
        dest="phmmer_dir",
        default=None,
        help="Folder containing PHMMER input files",
    )
    parser.add_argument(
        "-o",
        "--excel",
        dest="excel_output",
        default="FamilyB1_Seqs_final.xlsx",
        help="Excel output file path",
    )
    return parser.parse_args()


def read_tsv(tsv_path):
    if not os.path.exists(tsv_path):
        return []
    with open(tsv_path, newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        return [row for row in reader]


def extract_sequence(value):
    if not value:
        return None
    value = str(value)
    if value.startswith(">"):
        parts = value.split(" ", 1)
        return parts[1] if len(parts) > 1 else None
    return value


def parse_species_from_sequence(value):
    if not value:
        return None
    value = str(value)
    if value.startswith(">"):
        token = value[1:].split("_", 1)[0]
        return token or None
    return None


def collect_phmmer_files(args):
    candidates = []
    candidates.extend(args.inputs)

    if args.phmmer_dir and os.path.isdir(args.phmmer_dir):
        for name in sorted(os.listdir(args.phmmer_dir)):
            if name.lower().endswith((".txt", ".tsv")):
                candidates.append(os.path.join(args.phmmer_dir, name))
    else:
        default_dir = os.path.join(os.getcwd(), "phmmer_inputs")
        if os.path.isdir(default_dir):
            for name in sorted(os.listdir(default_dir)):
                if name.lower().endswith((".txt", ".tsv")):
                    candidates.append(os.path.join(default_dir, name))

    return [c for c in candidates if os.path.exists(c)]


def add_phmmer_sheet(wb, phmmer_files):
    ws = wb.create_sheet(title="PHMMER_RESULTS")
    ws.append(["source_file", "line"])

    if not phmmer_files:
        ws.append(["NA", "No PHMMER input files found"])
        return

    for path in phmmer_files:
        with open(path) as f:
            for line in f:
                if line.startswith("#") or not line.strip():
                    continue
                ws.append([os.path.basename(path), line.strip()])


def append_family_rows(ws, output_dirs, family_name, subfamilies):
    for subfamily in subfamilies:
        label = f"{subfamily} subfamily"
        first_row = True

        all_rows = []
        for output_dir in output_dirs:
            new_path = os.path.join(output_dir, family_name, subfamily, f"{subfamily}.tsv")
            old_path = os.path.join(output_dir, subfamily, f"{subfamily}.tsv")
            rows = read_tsv(new_path) or read_tsv(old_path)
            all_rows.extend(rows)

        if not all_rows:
            ws.append([label, subfamily, None, None, None, None, None, None, None])
            continue

        for row in all_rows:
            sequence = extract_sequence(row.get("protein_sequence"))
            species = parse_species_from_sequence(row.get("protein_sequence"))
            species = species.upper() if species else None
            ws.append([
                label if first_row else None,
                subfamily,
                row.get("accession"),
                row.get("tlen"),
                row.get("GeneID"),
                row.get("chromosome"),
                row.get("exon_count"),
                sequence,
                species,
            ])
            first_row = False


args = parse_args()

output_dirs = []
output_dirs.extend(args.output_dirs)
output_dirs.extend(resolve_output_dir(i, args.output_root) for i in args.inputs)

if args.output_root:
    for name in sorted(os.listdir(args.output_root)):
        candidate = os.path.join(args.output_root, name)
        if os.path.isdir(candidate) and name.startswith("FINAL_OUTPUT_"):
            output_dirs.append(candidate)
else:
    latest_root = find_latest_output_root()
    if latest_root:
        for name in sorted(os.listdir(latest_root)):
            candidate = os.path.join(latest_root, name)
            if os.path.isdir(candidate) and name.startswith("FINAL_OUTPUT_"):
                output_dirs.append(candidate)

if not output_dirs:
    output_dirs = [DEFAULT_OUTPUT_DIR]

print("🔄 Starting combined Excel generation...")
print(f"📊 Creating Excel workbook: {args.excel_output}")

wb = Workbook()
wb.remove(wb.active)

phmmer_files = collect_phmmer_files(args)
add_phmmer_sheet(wb, phmmer_files)

for family_name in GROUP_ORDER:
    ws = wb.create_sheet(title=family_name)
    ws.append([None] * 9)
    ws.append([
        None,
        "Gene",
        "Accession number",
        "protein length (aa)",
        "GeneID",
        "Chr",
        "exon",
        "Sequence",
        "Species",
    ])
    append_family_rows(ws, output_dirs, family_name, GROUP_SUBFAMILIES[family_name])

wb.save(args.excel_output)
print(f"\n✅ Excel saved: {args.excel_output}")
