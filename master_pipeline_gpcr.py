from Bio import Entrez, SeqIO
import time
from collections import defaultdict
from datetime import timedelta, datetime
import os
import io
import re
import argparse
import shutil
from openpyxl import Workbook

# ================= CONFIG =================
Entrez.email = "a71364@ualg.pt"

DEFAULT_INPUT_FILE = "Galaxy14_PHMMER_Takifugu_rubripes.txt"

SLEEP = 0.34
PROGRESS_STEP = 10
# =========================================


FAMILIES = {
    "SCTR": [
        "secretin receptor",
        "secretin",
        "sctr",
        "sctra",
        "sctrb",
    ],
    "GCGR": [
        "glucagon receptor",
        "gcgr",
        "gcgra",
        "gcgrb",
        "gcrpr",
    ],
    "GIPR": [
        "gastric inhibitory",
        "gipr",
        "gip receptor",
        "gipr2",
    ],
    "GLP1R": [
        "glucagon-like peptide-1",
        "glucagon like peptide 1",
        "glucagon-like peptide 1 receptor",
        "glp1r",
        "glp-1 receptor",
        "glpr1",
    ],
    "GLP2R": [
        "glucagon-like peptide-2",
        "glucagon like peptide 2",
        "glucagon-like peptide 2 receptor",
        "glp2r",
        "glp-2 receptor",
        "glpr2",
    ],
    "GHRHR": [
        "growth hormone releasing hormone receptor",
        "growth hormone releasing hormone receptor, like",
        "growth hormone-releasing",
        "ghrhr",
    ],
    "PTH1R": [
        "parathyroid hormone receptor 1",
        "parathyroid hormone/parathyroid hormone-related peptide receptor",
        "parathyroid hormone-related peptide receptor",
        "parathyroid hormone/parathyroid hormone-related peptide receptor-like",
        "pth1r",
        "pthr1",
        "pthr1r",
        "pth3r",
        "pthr3",
        "prpr",
        "prpra",
        "prprb",
        "phir",
        "pihr",
    ],
    "PTH2R": [
        "parathyroid hormone receptor 2",
        "parathyroid hormone 2 receptor",
        "pth2r",
        "pth2ra",
        "pth2rb",
        "pthr2",
        "pthr2r",
    ],
    "CRHR1": [
        "crhr1",
        "crfr1",
        "crh receptor 1",
        "crh-r1",
        "crf receptor 1",
        "crf1 receptor",
        "corticotropin releasing factor receptor 1",
        "corticotropin releasing hormone receptor 1",
        "corticotropin-releasing hormone receptor 1",
    ],
    "CRHR2": [
        "crhr2",
        "crfr2",
        "crh receptor 2",
        "crh-r2",
        "crf receptor 2",
        "crf2 receptor",
        "corticotropin releasing factor receptor 2",
        "corticotropin releasing hormone receptor 2",
        "corticotropin-releasing hormone receptor 2",
    ],
    "ADCYAP1R1": [
        "pacap",
        "adcyap1r1",
        "adcyap1r1a",
        "adcyap1r1b",
        "pituitary adenylate cyclase-activating polypeptide type i receptor",
        "pituitary adenylate cyclase-activating polypeptide type 1 receptor",
        "adenylate cyclase activating polypeptide 1",
        "pac1 receptor",
        "pacap type i receptor",
    ],
    "VIPR1": [
        "vasoactive intestinal peptide receptor 1",
        "vasoactive intestinal polypeptide receptor 1",
        # REMOVED: "vasoactive intestinal polypeptide receptor" (too generic - captures VIPR2!)
        "vipr1",
        "vipr1a",
        "vipr1b",
    ],
    "VIPR2": [
        "vasoactive intestinal peptide receptor 2",
        "vasoactive intestinal polypeptide receptor 2",
        "vipr2",
        "vipr2a",
        "vipr2b",
    ],
    "CALCR": [
        "calcitonin receptor",
        "calcr",
        "calcra",
        "calcrb",
        "calcr3",
    ],
    "CALCRL": [
        "calcitonin receptor-like",
        "calcitonin gene-related peptide type 1 receptor",
        "calcitonin gene-related peptide type 1 receptor-like",
        "calcitonin gene related peptide type 1 receptor",
        "calcrl",
        "calcrla",
        "calcrlb",
        "calcrl3",
    ],
}

CLASSIFICATION_ORDER = [
    "VIPR2",
    "VIPR1",
    "GLP2R",
    "GLP1R",
    "PTH2R",
    "PTH1R",
    "CRHR2",
    "CRHR1",
    "SCTR",
    "GCGR",
    "GIPR",
    "GHRHR",
    "ADCYAP1R1",
    "CALCR",
    "CALCRL",
]

GROUP_ORDER = ["CALCR", "CRHR", "GCGR", "SCTR", "PTHR"]

GROUP_SUBFAMILIES = {
    "CALCR": ["CALCRL", "CALCR"],
    "CRHR": ["CRHR1", "CRHR2"],
    "GCGR": ["GCGR", "GLP1R", "GLP2R", "GIPR"],
    "SCTR": ["GHRHR", "ADCYAP1R1", "VIPR1", "VIPR2", "SCTR"],
    "PTHR": ["PTH1R", "PTH2R"],
}

SUBFAMILY_TO_GROUP = {
    "CALCRL": "CALCR",
    "CALCR": "CALCR",
    "CRHR1": "CRHR",
    "CRHR2": "CRHR",
    "GCGR": "GCGR",
    "GLP1R": "GCGR",
    "GLP2R": "GCGR",
    "GIPR": "GCGR",
    "GHRHR": "SCTR",
    "ADCYAP1R1": "SCTR",
    "VIPR1": "SCTR",
    "VIPR2": "SCTR",
    "SCTR": "SCTR",
    "PTH1R": "PTHR",
    "PTH2R": "PTHR",
}

FAMILY_SHEETS = GROUP_ORDER + ["UNCLASSIFIED"]


def normalize_for_match(value):
    # Normalize punctuation/hyphen differences for robust keyword matching.
    return re.sub(r"[^a-z0-9]+", " ", str(value).lower()).strip()


NORMALIZED_FAMILY_KEYS = {
    subfamily: [normalize_for_match(k) for k in keys]
    for subfamily, keys in FAMILIES.items()
}


def is_np(acc):
    return acc.startswith("NP_")


def get_gene_id(acc):
    try:
        h = Entrez.elink(dbfrom="protein", db="gene", id=acc)
        r = Entrez.read(h)
        h.close()
        links = r[0].get("LinkSetDb", [])
        if links:
            return links[0]["Link"][0]["Id"]
    except:
        pass
    return None


def get_protein_sequence(acc):
    try:
        h = Entrez.efetch(db="protein", id=acc, rettype="fasta", retmode="text")
        seq = SeqIO.read(io.StringIO(h.read()), "fasta")
        h.close()
        return str(seq.seq)
    except:
        return "NA"


def get_species_abbrev(description_line):
    try:
        match = re.findall(r"\[([^\]]+)\]", description_line)
        if not match:
            return "NA"

        species = match[-1].strip()
        parts = [p for p in species.split() if p]
        if len(parts) < 2:
            return "NA"

        first = parts[0]
        last = parts[-1]
        abbrev = f"{first[0].upper()}{last[:2].lower()}"
        return abbrev
    except:
        return "NA"


def format_sequence_with_header(acc, sequence, description_line, subfamily=""):
    abbrev = get_species_abbrev(description_line)
    if sequence == "NA":
        return f">{abbrev}_{acc}_{subfamily} NA"
    return f">{abbrev}_{acc}_{subfamily} {sequence}"


def sanitize_name(value):
    safe = re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_")
    return safe or "output"


def resolve_output_dir(input_file, output_root):
    base = os.path.splitext(os.path.basename(input_file))[0]
    safe_base = sanitize_name(base)
    return os.path.join(output_root, f"FINAL_OUTPUT_{safe_base}")


def resolve_output_root(output_root=None):
    if output_root:
        return output_root
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(os.getcwd(), "outputs", f"run_{run_id}")


def normalize_sheet_title(value, used_titles):
    base = sanitize_name(value)[:31] or "Sheet"
    title = base
    suffix = 1
    while title in used_titles:
        trimmed = base[:28]
        title = f"{trimmed}_{suffix}"
        suffix += 1
    used_titles.add(title)
    return title


def read_phmmer_lines(input_file):
    lines = []
    with open(input_file) as f:
        for line in f:
            if line.startswith("#") or not line.strip():
                continue
            lines.append(line.strip())
    return lines


def build_row(record, subfamily, group):
    species = get_species_abbrev(record["line"])
    return {
        "order": record["order"],
        "family": group,
        "subfamily": subfamily,
        "accession": record["accession"],
        "tlen": record["tlen"],
        "gene_id": record.get("gene_id") or "NA",
        "chromosome": record.get("chromosome") or "NA",
        "exon_count": record.get("exon_count") or "NA",
        "sequence": record.get("sequence") or "NA",
        "species": species or "NA",
        "description": record["line"],
    }


def write_phmmer_sheet(wb, input_file, phmmer_lines):
    ws = wb.active
    ws.title = "PHMMER_RESULTS"
    ws.append(["source_file", "line"])
    if not phmmer_lines:
        ws.append([os.path.basename(input_file), "NA"])
        return
    for line in phmmer_lines:
        ws.append([os.path.basename(input_file), line])


def write_family_sheet(wb, family, selected_rows, rejected_rows):
    ws = wb.create_sheet(title=family)
    ws.append([
        "Subfamily",
        "Accession",
        "tlen",
        "GeneID",
        "chromosome",
        "exon_count",
        "Sequence",
        "Species",
        "Description",
    ])

    for row in selected_rows:
        ws.append([
            row["subfamily"],
            row["accession"],
            row["tlen"],
            row["gene_id"],
            row["chromosome"],
            row["exon_count"],
            row["sequence"],
            row["species"],
            row["description"],
        ])

    if rejected_rows:
        ws.append(["REJEITADAS ABAIXO", None, None, None, None, None, None, None, None])
        for row in rejected_rows:
            ws.append([
                row["subfamily"],
                row["accession"],
                row["tlen"],
                row["gene_id"],
                row["chromosome"],
                row["exon_count"],
                row["sequence"],
                row["species"],
                row["description"],
            ])


def write_species_excel(species_data, output_root):
    wb = Workbook()
    write_phmmer_sheet(wb, species_data["input_file"], species_data["phmmer_lines"])

    selected = species_data["selected"]
    rejected = species_data["rejected"]

    for family in GROUP_ORDER:
        selected_rows = [r for r in selected if r["family"] == family]
        rejected_rows = [r for r in rejected if r["family"] == family]
        write_family_sheet(wb, family, selected_rows, rejected_rows)

    unclassified_selected = [r for r in selected if r["family"] == "UNCLASSIFIED"]
    unclassified_rejected = [r for r in rejected if r["family"] == "UNCLASSIFIED"]
    write_family_sheet(wb, "UNCLASSIFIED", unclassified_selected, unclassified_rejected)

    output_path = os.path.join(output_root, f"{species_data['safe_base']}_FamilyB1.xlsx")
    wb.save(output_path)
    return output_path


def write_combined_workbook(all_species_data, output_root):
    wb = Workbook()
    wb.remove(wb.active)
    used_titles = set()

    for data in all_species_data:
        sheet_title = normalize_sheet_title(data["safe_base"], used_titles)
        ws = wb.create_sheet(title=sheet_title)
        ws.append([
            "Family",
            "Subfamily",
            "Accession",
            "tlen",
            "GeneID",
            "chromosome",
            "exon_count",
            "Sequence",
            "Species",
            "Description",
        ])

        selected = data["selected"]
        rejected = data["rejected"]

        for family in GROUP_ORDER:
            for subfamily in GROUP_SUBFAMILIES[family]:
                rows = [r for r in selected if r["subfamily"] == subfamily]
                if rows:
                    for row in rows:
                        ws.append([
                            family,
                            subfamily,
                            row["accession"],
                            row["tlen"],
                            row["gene_id"],
                            row["chromosome"],
                            row["exon_count"],
                            row["sequence"],
                            row["species"],
                            row["description"],
                        ])
                else:
                    ws.append([family, subfamily, None, None, None, None, None, None, None, None])

        unclassified_selected = [r for r in selected if r["family"] == "UNCLASSIFIED"]
        if unclassified_selected:
            for row in unclassified_selected:
                ws.append([
                    "UNCLASSIFIED",
                    "UNCLASSIFIED",
                    row["accession"],
                    row["tlen"],
                    row["gene_id"],
                    row["chromosome"],
                    row["exon_count"],
                    row["sequence"],
                    row["species"],
                    row["description"],
                ])
        else:
            ws.append(["UNCLASSIFIED", "UNCLASSIFIED", None, None, None, None, None, None, None, None])

        if rejected:
            ws.append(["REJEITADAS ABAIXO", None, None, None, None, None, None, None, None, None])
            for row in rejected:
                ws.append([
                    row["family"],
                    row["subfamily"],
                    row["accession"],
                    row["tlen"],
                    row["gene_id"],
                    row["chromosome"],
                    row["exon_count"],
                    row["sequence"],
                    row["species"],
                    row["description"],
                ])

    output_path = os.path.join(output_root, "All_species_combined.xlsx")
    wb.save(output_path)
    return output_path


def classify_subfamily(description_line):
    desc_norm = normalize_for_match(description_line)
    for subfamily in CLASSIFICATION_ORDER:
        keys = NORMALIZED_FAMILY_KEYS.get(subfamily, [])
        if any(k and k in desc_norm for k in keys):
            return subfamily
    return "UNCLASSIFIED"


def get_gene_summary(gene_id):
    try:
        h = Entrez.esummary(db="gene", id=gene_id, retmode="xml")
        summary = Entrez.read(h)
        h.close()

        docset = summary.get("DocumentSummarySet", {})
        docs = docset.get("DocumentSummary", [])
        if not docs:
            return {}

        doc = docs[0]
        genomic = (doc.get("GenomicInfo") or [{}])[0]

        exon_count = genomic.get("ExonCount")
        chromosome = doc.get("Chromosome") or genomic.get("ChrLoc")
        chr_accver = genomic.get("ChrAccVer")
        chr_start = genomic.get("ChrStart")
        chr_stop = genomic.get("ChrStop")

        return {
            "exon_count": exon_count,
            "chromosome": chromosome,
            "chr_accver": chr_accver,
            "chr_start": chr_start,
            "chr_stop": chr_stop,
        }
    except:
        return {}


def run_pipeline(input_file, output_root):
    records = []
    unique_accessions = []
    seen = set()
    phmmer_lines = []

    # ---------- READ INPUT ----------
    with open(input_file) as f:
        for line_number, line in enumerate(f, start=1):
            if line.startswith("#") or not line.strip():
                continue

            phmmer_lines.append(line.strip())
            parts = line.split()
            acc = parts[0]
            tlen = int(parts[2])

            records.append({
                "line": line.strip(),
                "accession": acc,
                "tlen": tlen,
                "order": line_number,
            })

            if acc not in seen:
                unique_accessions.append(acc)
                seen.add(acc)

    total = len(unique_accessions)

    gene_cache = {}
    seq_cache = {}
    gene_info_cache = {}

    start = time.time()
    next_progress = PROGRESS_STEP

    print(f"Total unique accessions: {total}\n")

    # ---------- FETCH GENE IDS + SEQUENCES ----------
    for i, acc in enumerate(unique_accessions, start=1):
        gene_cache[acc] = get_gene_id(acc)
        seq_cache[acc] = get_protein_sequence(acc)

        percent = (i / total) * 100
        elapsed = time.time() - start

        if percent >= next_progress or i == total:
            eta = (elapsed / i) * (total - i)
            print(
                f"[{int(percent)}%] {i}/{total} | "
                f"elapsed: {timedelta(seconds=int(elapsed))} | "
                f"ETA: {timedelta(seconds=int(eta))}"
            )
            next_progress += PROGRESS_STEP

        time.sleep(SLEEP)

    # ---------- FETCH GENE SUMMARY (EXON + CHR POSITION) ----------
    unique_gene_ids = sorted({gid for gid in gene_cache.values() if gid})
    gene_total = len(unique_gene_ids)
    start = time.time()
    next_progress = PROGRESS_STEP

    print(f"\nTotal unique GeneIDs: {gene_total}\n")

    for i, gid in enumerate(unique_gene_ids, start=1):
        gene_info_cache[gid] = get_gene_summary(gid)

        percent = (i / gene_total) * 100 if gene_total else 100
        elapsed = time.time() - start

        if percent >= next_progress or i == gene_total:
            eta = (elapsed / i) * (gene_total - i) if i else 0
            print(
                f"[Gene {int(percent)}%] {i}/{gene_total} | "
                f"elapsed: {timedelta(seconds=int(elapsed))} | "
                f"ETA: {timedelta(seconds=int(eta))}"
            )
            next_progress += PROGRESS_STEP

        time.sleep(SLEEP)

    # ---------- ATTACH DATA ----------
    for r in records:
        acc = r["accession"]
        r["gene_id"] = gene_cache.get(acc)
        r["sequence"] = seq_cache.get(acc)
        info = gene_info_cache.get(r["gene_id"], {})
        r["exon_count"] = info.get("exon_count") or "NA"
        r["chromosome"] = info.get("chromosome") or "NA"
        r["chr_accver"] = info.get("chr_accver") or "NA"
        r["chr_start"] = info.get("chr_start") or "NA"
        r["chr_stop"] = info.get("chr_stop") or "NA"

    # ---------- GROUP BY GENEID ----------
    by_gene = defaultdict(list)
    for r in records:
        key = r["gene_id"] or f"NO_GENE_{r['accession']}"
        by_gene[key].append(r)

    # ---------- FILTER (largest, NP > XP) ----------
    selected = []
    eliminated = []
    for _, group in by_gene.items():
        group.sort(
            key=lambda x: (x["tlen"], is_np(x["accession"])),
            reverse=True
        )
        selected.append(group[0])
        eliminated.extend(group[1:])

    selected_rows = []
    rejected_rows = []

    for r in sorted(selected, key=lambda x: x["order"]):
        subfamily = classify_subfamily(r["line"])
        group = SUBFAMILY_TO_GROUP.get(subfamily, "UNCLASSIFIED")
        if subfamily == "UNCLASSIFIED":
            group = "UNCLASSIFIED"
        selected_rows.append(build_row(r, subfamily, group))

    for r in sorted(eliminated, key=lambda x: x["order"]):
        subfamily = classify_subfamily(r["line"])
        group = SUBFAMILY_TO_GROUP.get(subfamily, "UNCLASSIFIED")
        if subfamily == "UNCLASSIFIED":
            group = "UNCLASSIFIED"
        rejected_rows.append(build_row(r, subfamily, group))

    print("\nPIPELINE COMPLETE ✅")

    safe_base = sanitize_name(os.path.splitext(os.path.basename(input_file))[0])
    return {
        "input_file": input_file,
        "safe_base": safe_base,
        "phmmer_lines": phmmer_lines,
        "selected": selected_rows,
        "rejected": rejected_rows,
    }


def parse_args():
    parser = argparse.ArgumentParser(
        description="GPCR B1 Family Pipeline - per-input outputs"
    )
    parser.add_argument(
        "-i",
        "--input",
        action="append",
        dest="inputs",
        default=[],
        help="PHMMER input file (can be used multiple times)",
    )
    parser.add_argument(
        "-d",
        "--input-dir",
        dest="input_dir",
        default=None,
        help="Directory with PHMMER input files to process",
    )
    parser.add_argument(
        "--output-root",
        dest="output_root",
        default=None,
        help="Root folder for all outputs (default: outputs/run_<timestamp>)",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    output_root = resolve_output_root(args.output_root)
    os.makedirs(output_root, exist_ok=True)
    print(f"\n=== Output root: {output_root} ===\n")
    input_files = list(args.inputs or [])
    if args.input_dir:
        for name in sorted(os.listdir(args.input_dir)):
            if name.lower().endswith((".txt", ".tsv")):
                input_files.append(os.path.join(args.input_dir, name))

    if not input_files:
        input_files = [DEFAULT_INPUT_FILE]
    all_species = []
    for input_file in input_files:
        print(f"\n=== Running pipeline for: {input_file} ===\n")
        species_data = run_pipeline(input_file, output_root)
        all_species.append(species_data)
        output_path = write_species_excel(species_data, output_root)
        print(f"Excel saved: {output_path}")

    combined_path = write_combined_workbook(all_species, output_root)
    print(f"\nCombined Excel saved: {combined_path}")
